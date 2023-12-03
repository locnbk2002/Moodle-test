import unittest
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import openpyxl


class TestAssignSystemRole(unittest.TestCase):
    def test_AssignSystemRole(self):
        self.workbook = openpyxl.load_workbook(filename="./testcase.xlsx")
        self.sheet = self.workbook["AssignSystemRole"]

        for row in range(2, self.sheet.max_row + 1):
            print("In testcase number " + str(row - 1))

            try:
                options = webdriver.ChromeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Chrome(options=options)
            except:
                options = webdriver.EdgeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Edge(options=options)

            self.driver.set_window_size(1460, 1080)
            self.driver.get("https://school.moodledemo.net/login/index.php")
            self.assertIn("Log in to the site | Mount Orange School", self.driver.title)

            # Find username and password
            username = self.driver.find_element(By.ID, "username")
            password = self.driver.find_element(By.ID, "password")

            # Enter credentials
            username.send_keys(self.sheet.cell(row=row, column=1).value)
            password.send_keys(self.sheet.cell(row=row, column=2).value)

            # Submit form
            submit = self.driver.find_element(By.ID, "loginbtn")
            submit.click()
            time.sleep(3)

            # Click on Site administration
            self.driver.find_element(By.LINK_TEXT, "Site administration").click()

            # Click on Users
            self.driver.find_element(By.LINK_TEXT, "Users").click()

            # Click on Assign system roles
            self.driver.find_element(By.LINK_TEXT, "Assign system roles").click()

            # Click on Course creator
            self.driver.find_element(By.LINK_TEXT, "Course creator").click()

            # Select users
            dropdown = self.driver.find_element(By.ID, "addselect")
            dropdown.find_element(
                By.XPATH,
                "//option[. = '" + self.sheet.cell(row=row, column=3).value + "']",
            ).click()
            self.driver.find_element(By.ID, "add").click()

            # Logout
            self.driver.find_element(By.CSS_SELECTOR, ".userpicture").click()
            self.driver.find_element(By.LINK_TEXT, "Log out").click()
            self.driver.quit()

    def tearDown(self):
        print(f"\n{self._testMethodName} finished\n")


if __name__ == "__main__":
    unittest.main()
