import unittest
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl


class TestRemoveSystemRoleBonus(unittest.TestCase):
    def test_RemoveSystemRole(self):
        self.workbook = openpyxl.load_workbook(filename="./testcase.xlsx")
        self.sheet = self.workbook["AssignSystemRoleBonus"]

        for row in range(2, self.sheet.max_row + 1):
            print("In testcase number " + str(row - 1))
            get_value = lambda column: self.sheet.cell(row=row, column=column).value

            try:
                options = webdriver.ChromeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Chrome(options=options)
            except:
                options = webdriver.EdgeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Edge(options=options)

            self.driver.set_window_size(1460, 1080)
            self.driver.get(get_value(1))
            self.assertIn(get_value(2), self.driver.title)

            # Find username and password
            username = self.driver.find_element(By.ID, get_value(3))
            password = self.driver.find_element(By.ID, get_value(4))

            # Enter credentials
            username.send_keys(get_value(5))
            password.send_keys(get_value(6))

            # Submit form
            submit = self.driver.find_element(By.ID, get_value(7))
            submit.click()
            time.sleep(3)

            # Click on Site administration
            self.driver.find_element(By.LINK_TEXT, get_value(8)).click()

            # Click on Users
            self.driver.find_element(By.LINK_TEXT, get_value(9)).click()

            # Click on Assign system roles
            self.driver.find_element(By.LINK_TEXT, get_value(10)).click()

            # Click on Course creator
            self.driver.find_element(By.LINK_TEXT, get_value(11)).click()

            # Select users
            dropdown = self.driver.find_element(By.ID, get_value(13))
            dropdown.find_element(By.XPATH, get_value(14)).click()
            self.driver.find_element(By.ID, get_value(16)).click()

            # Logout
            self.driver.find_element(By.CSS_SELECTOR, get_value(17)).click()
            self.driver.find_element(By.LINK_TEXT, get_value(18)).click()
            self.driver.quit()

    def tearDown(self):
        print(f"\n{self._testMethodName} finished\n")


if __name__ == "__main__":
    unittest.main()
